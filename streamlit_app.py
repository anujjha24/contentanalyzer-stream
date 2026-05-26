"""
Content Analyzer — Streamlit host with FastAPI backend.

Architecture:
  • FastAPI runs on a background thread (port 8502 by default).
  • Streamlit serves the page via components.html(), injecting the API base URL.
  • All original HTML/CSS/JS is preserved; only /static/* paths are rewritten.
  • MongoDB logic is unchanged; subprocess is eliminated.
"""

import io
import json
import logging
import os
import re
import threading
import zipfile
from datetime import datetime
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
import uvicorn
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).resolve().parent
STATIC_DIR   = BASE_DIR / "static"
TEMPLATE_FILE = BASE_DIR / "brand_comparison_template.xlsx"
HTML_FILE    = BASE_DIR / "content_analyzer.html"

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
)
logger = logging.getLogger("content_analyzer")

# ── FastAPI app ───────────────────────────────────────────────────────────────
api = FastAPI(title="Content Analyzer API", docs_url=None, redoc_url=None)
api.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
api.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

API_PORT = int(os.getenv("API_PORT", "8502"))


# ── Helpers ───────────────────────────────────────────────────────────────────
def to_json_safe(value):
    from utils.helpers import to_json_safe as _safe
    return _safe(value)


def ok(data) -> JSONResponse:
    return JSONResponse(content=to_json_safe(data))


def err(msg: str, status: int = 500) -> JSONResponse:
    return JSONResponse(content={"success": False, "error": str(msg)}, status_code=status)


def xlsx_response(data: bytes, filename: str) -> Response:
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Routes ────────────────────────────────────────────────────────────────────

@api.get("/healthz")
def healthz():
    return {"status": "ok"}


@api.get("/api/channels-dates")
def get_channels_dates():
    try:
        from utils.db import get_collections
        from pymongo import ASCENDING
        processed_files, _, _ = get_collections()
        docs = processed_files.find(
            {}, {"_id": 0, "channel_name": 1, "date": 1}
        ).sort([("channel_name", ASCENDING), ("date", ASCENDING)])
        return ok([{"channel_name": d.get("channel_name", ""), "date": d.get("date", "")} for d in docs])
    except Exception as exc:
        logger.exception("channels-dates failed: %s", exc)
        return err(str(exc))


@api.get("/api/dashboard")
def get_dashboard(channel: str = "", date: str = "", source: str = "TABSONS-BARC", data_type: str = "COUNT"):
    if not channel or not date:
        return err("channel and date required", 400)
    try:
        from utils.processing import get_dashboard_data
        result = get_dashboard_data(channel, date, source, data_type)
        return ok(result)
    except ValueError as exc:
        return err(str(exc), 404)
    except Exception as exc:
        logger.exception("dashboard failed: %s", exc)
        return err(str(exc))


@api.get("/api/sheets")
def get_sheets(channel: str = "", date: str = ""):
    if not channel or not date:
        return err("channel and date required", 400)
    try:
        from utils.db import get_collections
        from pymongo import ASCENDING
        _, sheets, _ = get_collections()
        docs = sheets.find(
            {"channel_name": channel, "date": date},
            {"_id": 0, "sheet_name": 1, "row_count": 1, "col_count": 1},
        ).sort([("_id", ASCENDING)])
        return ok([{
            "sheet_name": d.get("sheet_name", ""),
            "row_count": d.get("row_count", 0),
            "col_count": d.get("col_count", 0),
        } for d in docs])
    except Exception as exc:
        logger.exception("sheets failed: %s", exc)
        return err(str(exc))


@api.get("/api/sheet-data")
def get_sheet_data(channel: str = "", date: str = "", sheet: str = ""):
    if not channel or not date or not sheet:
        return err("channel, date, and sheet required", 400)
    try:
        from utils.db import get_collections
        _, sheets, _ = get_collections()
        row = sheets.find_one(
            {"channel_name": channel, "date": date, "sheet_name": sheet},
            {"_id": 0, "file_id": 1, "channel_name": 1, "date": 1,
             "sheet_name": 1, "headers": 1, "rows": 1, "row_count": 1,
             "col_count": 1, "uploaded_at": 1},
        )
        if not row:
            return err("Sheet not found", 404)
        return ok(row)
    except Exception as exc:
        logger.exception("sheet-data failed: %s", exc)
        return err(str(exc))


@api.get("/api/commercial-comparison")
def get_commercial_comparison(channel: str = "", date: str = ""):
    if not channel or not date:
        return err("channel and date required", 400)
    try:
        from utils.db import get_collections
        from pymongo import ASCENDING
        from utils.helpers import parse_count_cell
        _, sheets, brand_modifications = get_collections()
        row = sheets.find_one(
            {"channel_name": channel, "date": date, "sheet_name": "COMMERCIAL COMPARISION"},
            {"_id": 0, "rows": 1},
        )
        mods = list(brand_modifications.find(
            {"channel_name": channel, "date": date}, {"_id": 0}
        ).sort([("timestamp", ASCENDING)]))
    except Exception as exc:
        logger.exception("commercial-comparison DB failed: %s", exc)
        return err(str(exc))

    if not row:
        return err("Commercial comparison data not found", 404)

    rows = row.get("rows") or []
    headers_row = None
    matched_rows, unmatched_rows = [], []
    section = "none"

    for r in rows:
        first_cell = str(r[0]).strip() if r else ""
        if first_cell == "SOURCE" and headers_row is None:
            headers_row = r
            continue
        if "BARC COMMERCIAL vs NCT COMMERCIAL" in first_cell and "MATCHED" in first_cell:
            section = "matched"; continue
        if "NCT COMMERCIAL BRANDS" in first_cell and "NOT MATCHED" in first_cell:
            section = "unmatched"; continue
        if first_cell == "MATCHED" or ("MATCHED" in first_cell and "UNMATCHED" not in first_cell
                and "NCT COMMERCIAL BRANDS" not in first_cell and "BARC COMMERCIAL" not in first_cell
                and first_cell not in ("MATCHING COMMERCIAL TOTAL",)):
            section = "matched"; continue
        if "NOT MATCHED" in first_cell or "UNMATCHED" in first_cell:
            section = "unmatched"; continue
        if first_cell in ("", "MATCHING COMMERCIAL TOTAL", "NCT UNMATCHED TOTAL", "GRAND TOTAL") \
                or "COMMERCIAL COMPARISION" in first_cell:
            continue

        if headers_row and len(r) >= 5:
            labels = ["source","channel_name","date","barc_brand","nct_brand",
                      "barc_count","nct_count","barc_duration","nct_duration",
                      "nct_ps_count","nct_ps_duration","remarks"]
            row_dict = {label: (str(r[j]) if j < len(r) and r[j] is not None else "")
                        for j, label in enumerate(labels)}
            if section == "matched" and first_cell in ("BARC XML", "NCT"):
                matched_rows.append(row_dict)
            elif section == "unmatched" and first_cell in ("NCT", "BARC XML"):
                unmatched_rows.append(row_dict)

    for mod in mods:
        action = mod.get("action")
        brand_name = mod.get("brand_name", "")
        if action == "remove_from_matched":
            removed = None
            for j, rd in enumerate(matched_rows):
                if rd.get("barc_brand", "").strip() == brand_name.strip():
                    removed = matched_rows.pop(j); break
            if not removed:
                for j, rd in enumerate(matched_rows):
                    if rd.get("nct_brand", "").strip() == brand_name.strip():
                        removed = matched_rows.pop(j); break
            if removed:
                removed["remarks"] = "REMOVED FROM MATCHED"
                unmatched_rows.append(removed)
        elif action == "merge_to_matched":
            target_barc = mod.get("target_barc_brand") or ""
            merged = None
            for j, rd in enumerate(unmatched_rows):
                if rd.get("nct_brand", "").strip() == brand_name.strip():
                    merged = unmatched_rows.pop(j); break
            if merged and target_barc:
                for rd in matched_rows:
                    if rd.get("barc_brand", "").strip() == target_barc.strip():
                        existing = parse_count_cell(rd.get("nct_count", "0"))
                        merge_v  = parse_count_cell(merged.get("nct_count", "0"))
                        rd["nct_count"] = str(existing + merge_v)
                        rd["remarks"] = "MATCHED (MERGED)"
                        break

    return ok({"matched": matched_rows, "unmatched": unmatched_rows, "headers": headers_row or []})


@api.post("/api/commercial/move-brand")
async def move_brand(request: Request):
    data = await request.json()
    channel     = data.get("channel", "")
    date        = data.get("date", "")
    action      = data.get("action", "")
    brand_name  = data.get("brand_name", "")
    target_barc = data.get("target_barc_brand", "")

    if not all([channel, date, action, brand_name]):
        return err("Missing required fields", 400)

    try:
        from utils.db import get_collections
        from utils.helpers import utc_now_iso
        _, _, brand_modifications = get_collections()
        brand_modifications.insert_one({
            "channel_name": channel, "date": date,
            "action": action, "brand_name": brand_name,
            "target_barc_brand": target_barc,
            "timestamp": utc_now_iso(),
        })
    except Exception as exc:
        logger.exception("move-brand failed: %s", exc)
        return err(str(exc))

    return ok({"success": True, "message": f"Brand '{brand_name}' {action} successfully"})


@api.post("/api/commercial/undo-modifications")
async def undo_modifications(request: Request):
    data = await request.json()
    channel = data.get("channel", "")
    date    = data.get("date", "")
    try:
        from utils.db import get_collections
        _, _, brand_modifications = get_collections()
        brand_modifications.delete_many({"channel_name": channel, "date": date})
    except Exception as exc:
        logger.exception("undo-modifications failed: %s", exc)
        return err(str(exc))
    return ok({"success": True})


@api.post("/api/compare")
async def compare_report(file: UploadFile = File(...)):
    file_bytes = await file.read()
    try:
        from utils.processing import run_comparison, upload_to_db
        xlsx_bytes, output_filename, stats = run_comparison(file_bytes, file.filename)
        channel = stats.get("channel", "UNKNOWN")
        date    = stats.get("date", "00/00/0000")
        try:
            upload_to_db(xlsx_bytes, channel, date, output_filename)
        except Exception as db_err:
            logger.warning("MongoDB upload failed (non-fatal): %s", db_err)
        return xlsx_response(xlsx_bytes, output_filename)
    except Exception as exc:
        logger.exception("compare failed: %s", exc)
        return err(str(exc))


@api.post("/analyze")
async def analyze(request: Request):
    """Batch file processing endpoint."""
    form = await request.form()
    files = form.getlist("files")
    if not files:
        return err("No files provided", 400)

    from utils.processing import run_comparison, upload_to_db
    results, errors = [], []

    for uploaded_file in files:
        try:
            file_bytes = await uploaded_file.read()
            xlsx_bytes, output_filename, stats = run_comparison(file_bytes, uploaded_file.filename)
            channel = stats.get("channel", "UNKNOWN")
            date    = stats.get("date", "00/00/0000")
            try:
                upload_to_db(xlsx_bytes, channel, date, output_filename)
            except Exception as db_err:
                logger.warning("MongoDB upload failed: %s", db_err)
            results.append({"fname": output_filename, "data": xlsx_bytes})
        except Exception as exc:
            errors.append({"file": uploaded_file.filename, "error": str(exc)})

    if len(files) == 1 and len(results) == 1:
        r = results[0]
        return xlsx_response(r["data"], r["fname"])

    if results:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in results:
                zf.writestr(r["fname"], r["data"])
        zip_buf.seek(0)
        zip_name = f"barc_nct_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return Response(
            content=zip_buf.read(),
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{zip_name}"'},
        )

    return err("All files failed", 500)


@api.get("/api/download/preprocessed")
def download_preprocessed(channel: str = "", date: str = ""):
    try:
        from utils.db import get_collections
        from utils.helpers import workbook_bytes_from_document
        processed_files, _, _ = get_collections()
        row = processed_files.find_one(
            {"channel_name": channel, "date": date},
            {"_id": 0, "xlsx_data": 1, "original_filename": 1},
        )
    except Exception as exc:
        return err(str(exc))

    if not row:
        return err("File not found", 404)

    return xlsx_response(
        workbook_bytes_from_document(row),
        row.get("original_filename") or "report.xlsx",
    )


@api.get("/api/download/updated")
def download_updated(channel: str = "", date: str = ""):
    try:
        from utils.db import get_collections
        from utils.helpers import workbook_bytes_from_document
        from pymongo import ASCENDING
        processed_files, _, brand_modifications = get_collections()
        row = processed_files.find_one(
            {"channel_name": channel, "date": date},
            {"_id": 0, "xlsx_data": 1, "original_filename": 1},
        )
        mods = list(brand_modifications.find(
            {"channel_name": channel, "date": date}, {"_id": 0}
        ).sort([("timestamp", ASCENDING)]))
    except Exception as exc:
        return err(str(exc))

    if not row:
        return err("File not found", 404)

    xlsx_data = workbook_bytes_from_document(row)

    if mods:
        import openpyxl
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_data))
        if "COMMERCIAL COMPARISION" in workbook.sheetnames:
            ws = workbook["COMMERCIAL COMPARISION"]
            last_row = ws.max_row + 2
            ws.cell(row=last_row, column=1, value="MODIFICATIONS APPLIED:")
            for i, mod in enumerate(mods):
                ws.cell(row=last_row+i+1, column=1,
                        value=f"{mod.get('action')}: {mod.get('brand_name')}")
                if mod.get("target_barc_brand"):
                    ws.cell(row=last_row+i+1, column=2,
                            value=f"\u2192 {mod.get('target_barc_brand')}")
        buf = io.BytesIO()
        workbook.save(buf)
        xlsx_data = buf.getvalue()

    filename = row.get("original_filename") or "report.xlsx"
    if mods:
        base, ext = os.path.splitext(filename)
        filename = f"{base}_UPDATED{ext}"

    return xlsx_response(xlsx_data, filename)


@api.get("/api/template")
def download_template():
    if not TEMPLATE_FILE.exists():
        return err("Template file not found", 404)
    return xlsx_response(TEMPLATE_FILE.read_bytes(), "brand_comparison_template.xlsx")


# ── Background API server ─────────────────────────────────────────────────────

_api_started = False
_api_lock = threading.Lock()


def _start_api_server():
    global _api_started
    with _api_lock:
        if _api_started:
            return
        _api_started = True

    def _run():
        uvicorn.run(api, host="0.0.0.0", port=API_PORT, log_level="warning")
    if "api_started" not in st.session_state:
        t = threading.Thread(target=_run, daemon=True)
        t.start()
        st.session_state.api_started = True
        logger.info("FastAPI backend started on port %d", API_PORT)


# ── HTML preparation ──────────────────────────────────────────────────────────

def _prepare_html(api_base: str) -> str:
    """Load content_analyzer.html and rewrite static asset paths."""
    html = HTML_FILE.read_text(encoding="utf-8")

    # Rewrite /static/... references to point at the API server
    html = html.replace(
        'href="/static/styles.css"',
        f'href="{api_base}/static/styles.css"',
    )
    html = html.replace(
        'src="/static/app.js"',
        f'src="{api_base}/static/app.js"',
    )

    # Inject API base URL so app.js can resolve fetch() calls correctly
    injection = f"""
<script>
  window.STREAMLIT_API_BASE = "{api_base}";
</script>
"""
    html = html.replace("</head>", injection + "</head>", 1)
    return html


# ── Streamlit page ────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Content Analyzer",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    # Hide Streamlit chrome so the custom UI fills the whole viewport
    st.markdown(
        """
        <style>
        #MainMenu, header, footer { visibility: hidden; }
        .stApp { margin: 0; padding: 0; }
        .block-container { padding: 0 !important; max-width: 100% !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Start the FastAPI backend (idempotent)
    _start_api_server()

    api_base = f"http://localhost:{API_PORT}"
    html_content = _prepare_html(api_base)

    components.html(
        html_content,
        height=900,
        scrolling=True,
    )


if __name__ == "__main__":
    main()
