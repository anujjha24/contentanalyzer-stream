"""
BARC vs NCT Comparison System  — v3 (PromoSponsor column approach)
Input : brand_comparison_template.xlsx
Output: barc_nct_comparison.xlsx

KEY DESIGN CHANGE (v3):
  Instead of prefixing "PROMO SPONSOR-" into the NCT brand field, we now add a
  SEPARATE column called "NCT PromoSponsor" in the COMPARISON sheet.

  For every NCT row that matches a BARC Sponsorship Promo brand, the matched
  BARC brand name is written into "NCT PromoSponsor" (e.g. "PATANJALI HONEY").
  The original "NCT brand" field is left completely untouched.

  This means:
    - A row can be COMMERCIAL in NCT Program Type AND also have a PromoSponsor
      label — which is exactly what happens in real data (e.g. PRAYAG PTMT
      FAUCETS has 26 pure commercial rows + 1 promo-sponsor row).
    - In the COMMERCIAL COMPARISON sheet, for each brand the table shows:
        BARC commercial count/dur | NCT commercial count/dur |
        NCT PS count/dur (rows for that brand tagged as PS) |
        REMARKS
      The NCT commercial count/dur EXCLUDES the PS-tagged rows for that brand.
    - Summary Table 1 NCT COMMERCIAL totals also exclude all PS-matched rows.
    - Summary Table 2 shows per-brand PS detail sourced from ps_brand_nct_map.

Sheets:
  1. COMPARISON          — original cols + "NCT PromoSponsor" + "Duration +-" + "Remarks"
  2. SUMMARY             — Table 1 (content type totals), Table 2 (PS brand detail),
                           Table 3 (unique brand counts), Commercial Match Summary
  3. COMMERCIAL COMPARISION
  4. DETAILED ANALYSIS
"""

import re
from datetime import timedelta
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurable ──────────────────────────────────────────────────────────────
INPUT_FILE           = "brand_comparison_template.xlsx"
OUTPUT_FILE          = "barc_nct_comparison.xlsx"
SIMILARITY_THRESHOLD = 0.80
TIME_TOLERANCE_SECS  = 1
# ─────────────────────────────────────────────────────────────────────────────

INVALID_BRAND_NAMES = {"COMMERCIAL", "PROMO", "PROGRAM", "STORY BLOCK"}

# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_str(val) -> str:
    if val is None: return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s

def hms_to_secs(val) -> int:
    if val is None or (isinstance(val, float) and pd.isna(val)): return 0
    if isinstance(val, timedelta): return int(val.total_seconds())
    s = safe_str(val)
    if not s: return 0
    m = re.match(r"^(\d+):(\d{2}):(\d{2})$", s)
    if m: return int(m.group(1))*3600 + int(m.group(2))*60 + int(m.group(3))
    try: return int(float(s))
    except: return 0

def secs_to_hms(secs: int) -> str:
    neg  = secs < 0
    secs = abs(int(secs))
    h, r = divmod(secs, 3600)
    m, s = divmod(r, 60)
    return f"{'−' if neg else '+'}{h:02d}:{m:02d}:{s:02d}"

def fmt_hms(secs: int) -> str:
    secs = abs(int(secs))
    h, r = divmod(secs, 3600)
    m, s = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"

def normalise_type(val: str) -> str:
    v = val.strip().upper()
    return "PROGRAM" if v == "STORY BLOCK" else v

def normalise_for_match(s: str) -> str:
    s = s.upper().strip()
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def smart_match(brand: str, nct_val: str) -> float:
    a = normalise_for_match(brand)
    b = normalise_for_match(nct_val)
    if not a or not b: return 0.0
    pattern = r'\b' + r'\s+'.join(re.escape(w) for w in a.split()) + r'\b'
    if re.search(pattern, b): return 1.0
    sig = [w for w in a.split() if len(w) > 2]
    if sig and re.search(r'\b' + re.escape(sig[0]) + r'\b', b): return 0.85
    return SequenceMatcher(None, a, b).ratio()

def parse_sponsor_brands(raw: str) -> list:
    """Parse BARC PromoSponsorName field → list of (brand_str, secs)."""
    out = []
    if not raw: return out
    for entry in raw.split(","):
        entry = entry.strip().strip("[]")
        if not entry: continue
        parts = entry.split("|")
        if len(parts) < 2: continue
        brand = parts[0].strip()
        if brand: out.append((brand, hms_to_secs(parts[1].strip())))
    return out

def find_brand_in_nct(brand: str, nct_rows: pd.DataFrame, threshold: float):
    best_score, best_idx, best_field = 0, None, None
    for idx, row in nct_rows.iterrows():
        for field in ["NCT brand", "NCT story"]:
            val = safe_str(row.get(field, ""))
            if not val: continue
            score = smart_match(brand, val)
            if score > best_score:
                best_score, best_idx, best_field = score, idx, field
    if best_score >= threshold:
        return best_idx, best_score, best_field
    return None, best_score, None

def normalise_name(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", s.upper().strip())

def best_match_in_list(query: str, candidates: list, threshold: float):
    norm_q = normalise_name(query)
    if not norm_q: return None, 0.0
    best_score, best_cand = 0.0, None
    for cand in candidates:
        norm_c = normalise_name(cand)
        if not norm_c: continue
        if norm_q == norm_c:                           score = 1.0
        elif norm_q in norm_c or norm_c in norm_q:    score = 0.9
        else:                                          score = smart_match(query, cand)
        if score > best_score:
            best_score, best_cand = score, cand
    return (best_cand, best_score) if best_score >= threshold else (None, best_score)

# ── Styling ───────────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
OK_FILL    = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL  = PatternFill("solid", fgColor="FFEB9C")
ERR_FILL   = PatternFill("solid", fgColor="FFC7CE")
TITLE_FILL = PatternFill("solid", fgColor="2E75B6")
SEC_FILL   = PatternFill("solid", fgColor="BDD7EE")
BARC_FILL  = PatternFill("solid", fgColor="DEEAF1")
NCT_FILL   = PatternFill("solid", fgColor="E2EFDA")
PS_FILL    = PatternFill("solid", fgColor="EDE7F6")   # soft purple for PS column
CONC_FILL  = PatternFill("solid", fgColor="FFF2CC")
MATCH_FILL = PatternFill("solid", fgColor="E2EFDA")
MISS_FILL  = PatternFill("solid", fgColor="FFC7CE")

WHITE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
TITLE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CONC_FONT   = Font(name="Arial", bold=True, size=10, color="7B3F00")
PS_FONT     = Font(name="Arial", bold=True, size=10, color="4A148C")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN        = Side(style="thin",   color="BFBFBF")
MED         = Side(style="medium", color="595959")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

def sc(ws, row, col, val, font=None, fill=None, align=CENTER, border=THIN_BORDER):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = font   or NORMAL_FONT
    c.fill      = fill   or PatternFill()
    c.alignment = align
    c.border    = border
    return c

def hdr(ws, row, col, val, fill=HDR_FILL):
    return sc(ws, row, col, val, font=WHITE_FONT, fill=fill, align=CENTER)

def merge_title(ws, row, c1, c2, val, fill=TITLE_FILL):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=val)
    c.font = TITLE_FONT; c.fill = fill; c.alignment = CENTER; c.border = MED_BORDER
    return c

def generate_conclusion(label: str, barc_secs: int, nct_secs: int) -> str:
    if nct_secs == barc_secs:
        return f"NCT AND BARC {label.upper()} DURATION MATCHES PERFECTLY"
    diff = abs(nct_secs - barc_secs)
    if nct_secs > barc_secs:
        return f"NCT TAGGING MORE {label.upper()} THAN BARC BY {fmt_hms(diff)}"
    return f"NCT TAGGING LESS {label.upper()} THAN BARC BY {fmt_hms(diff)}"

# ── Core comparison ───────────────────────────────────────────────────────────

def compare_rows(df_barc, df_nct, threshold, tol):
    """
    Walk every BARC row, find matching NCT window, compute duration diff and remarks.

    Returns
    -------
    barc_dur_pm      : list[str]   duration +/- string per BARC row
    barc_remarks     : list[str]   remark per BARC row
    ps_matched_nct   : set[int]    NCT positional indices matched as PromoSponsor
    nct_ps_label     : dict[int, str]
                       {nct_idx -> barc_brand_name}
                       Written into the new "NCT PromoSponsor" column.
    ps_brand_nct_map : dict[str, {"count": int, "secs": int}]
                       Per BARC brand: how many distinct NCT rows matched + their total duration.
    """
    nct = df_nct.copy().reset_index(drop=True)
    nct["_ss"] = nct["TelecastStartTime"].apply(hms_to_secs)

    barc_dur_pm, barc_remarks = [], []
    ps_matched_nct  = set()            # NCT positional indices claimed as PS
    nct_ps_label    = {}               # nct_idx -> BARC brand name (for new column)
    ps_brand_nct_idx = {}              # brand -> set of nct_idx  (deduped globally)

    for _, brow in df_barc.iterrows():
        b_start  = hms_to_secs(brow["TelecastStartTime"])
        b_end    = hms_to_secs(brow["TelecastEndTime"])
        b_dur    = hms_to_secs(brow["TelecastDuration"])
        b_ctype  = normalise_type(safe_str(brow["BARC ContentType"]))
        b_title  = safe_str(brow["BARC Title"])
        b_spname = safe_str(brow["BARC PromoSponsorName"])

        mask    = (nct["_ss"] >= (b_start - tol)) & (nct["_ss"] < (b_end + tol))
        nct_win = nct[mask].copy()

        nct_types = nct_win["NCT Program Type"].apply(
                        lambda v: normalise_type(safe_str(v))).unique().tolist()
        nct_total = sum(hms_to_secs(v) for v in nct_win["TelecastDuration"])
        dur_diff  = nct_total - b_dur
        dur_str   = secs_to_hms(dur_diff) if dur_diff != 0 else ""

        remarks = []
        is_sp   = (b_ctype == "PROMO" and
                   b_title.lower().startswith("sponsorship promo"))

        if not is_sp and len(nct_win) > 0:
            mismatched = [t for t in nct_types if t != b_ctype]
            if mismatched and len(mismatched) == len(nct_types):
                remarks.append("WRONG NCT PROGRAM TYPE")

        if is_sp and b_spname:
            for brand, brand_secs in parse_sponsor_brands(b_spname):
                found_idx, _, _ = find_brand_in_nct(brand, nct_win, threshold)

                if found_idx is None:
                    remarks.append(f"{brand} MISSING IN NCT")
                else:
                    # ── Record as PS-matched ──────────────────────────────────
                    ps_matched_nct.add(found_idx)
                    # Store the BARC brand name for the new PS column
                    # If multiple brands point to the same NCT row, concatenate
                    existing = nct_ps_label.get(found_idx, "")
                    nct_ps_label[found_idx] = (
                        existing + ", " + brand if existing else brand
                    )
                    # Track globally deduped NCT indices per brand
                    ps_brand_nct_idx.setdefault(brand, set()).add(found_idx)

                    nct_bd = hms_to_secs(safe_str(nct_win.loc[found_idx, "TelecastDuration"]))
                    if nct_bd != brand_secs:
                        remarks.append("BRAND DURATION NOT MATCHING")

        if not is_sp and dur_diff != 0:
            diff_types = [t for t in nct_types if t != b_ctype]
            if b_ctype == "COMMERCIAL":
                for nt in nct_types:
                    remarks.append(f"{nt} DURATION NOT MATCHING {secs_to_hms(dur_diff)}")
            else:
                remark = f"DURATION DIFF {secs_to_hms(dur_diff)}"
                if diff_types:
                    remark += f" | DIFFERENT NCT PROGRAM TYPE FOR {b_ctype}"
                remarks.append(remark)

        barc_dur_pm.append(dur_str)
        barc_remarks.append(" | ".join(remarks) if remarks else "OK")

    # Build ps_brand_nct_map: sum actual durations of deduped matched NCT rows per brand
    ps_brand_nct_map = {}
    for brand, idx_set in ps_brand_nct_idx.items():
        total_secs = sum(
            hms_to_secs(safe_str(nct.loc[i, "TelecastDuration"]))
            for i in idx_set if i in nct.index
        )
        ps_brand_nct_map[brand] = {"count": len(idx_set), "secs": total_secs}

    return barc_dur_pm, barc_remarks, ps_matched_nct, nct_ps_label, ps_brand_nct_map

# ── Brand analysis ────────────────────────────────────────────────────────────

def analyse_brands(df_barc, ps_brand_nct_map, threshold):
    """
    Build BARC PromoSponsor brand totals and match them to ps_brand_nct_map.
    Uses exact normalised key lookup first, then single-best fuzzy fallback —
    never accumulates multiple keys for one BARC brand.
    """
    barc_brands = {}
    for _, row in df_barc.iterrows():
        for brand, secs in parse_sponsor_brands(safe_str(row.get("BARC PromoSponsorName", ""))):
            entry = barc_brands.setdefault(brand, {"secs": 0, "count": 0})
            entry["secs"]  += secs
            entry["count"] += 1

    norm_ps_map = {normalise_for_match(k): (k, v) for k, v in ps_brand_nct_map.items()}

    brand_matches = []
    for barc_b, barc_info in sorted(barc_brands.items()):
        norm_b = normalise_for_match(barc_b)

        # 1) Exact normalised key match
        if norm_b in norm_ps_map:
            ps_key, ps_info = norm_ps_map[norm_b]
            brand_matches.append({
                "barc_brand": barc_b, "barc_secs": barc_info["secs"],
                "barc_count": barc_info["count"],
                "nct_brand":  ps_key,  "nct_secs": ps_info["secs"],
                "nct_count":  ps_info["count"],
                "score": 1.0, "matched": True,
                "dur_diff": ps_info["secs"] - barc_info["secs"],
            })
            continue

        # 2) Single-best fuzzy fallback
        best_score, best_ps_key, best_ps_info = 0.0, None, None
        for ps_b, ps_info in ps_brand_nct_map.items():
            score = smart_match(barc_b, ps_b)
            if score >= threshold and score > best_score:
                best_score, best_ps_key, best_ps_info = score, ps_b, ps_info

        if best_ps_key:
            brand_matches.append({
                "barc_brand": barc_b, "barc_secs": barc_info["secs"],
                "barc_count": barc_info["count"],
                "nct_brand":  best_ps_key, "nct_secs": best_ps_info["secs"],
                "nct_count":  best_ps_info["count"],
                "score": round(best_score, 2), "matched": True,
                "dur_diff": best_ps_info["secs"] - barc_info["secs"],
            })
        else:
            brand_matches.append({
                "barc_brand": barc_b, "barc_secs": barc_info["secs"],
                "barc_count": barc_info["count"],
                "nct_brand": "— NOT FOUND —", "nct_secs": 0, "nct_count": 0,
                "score": round(best_score, 2), "matched": False, "dur_diff": None,
            })

    return barc_brands, brand_matches

# ── COMPARISON sheet ──────────────────────────────────────────────────────────

def build_comparison_sheet(ws, df_orig, barc_dur_pm, barc_remarks, nct_ps_label):
    """
    New column layout:
      original_cols... | NCT PromoSponsor | Duration +- | Remarks
    NCT PromoSponsor is filled only for NCT rows that were matched as PS.
    The original NCT brand field is NEVER modified.
    """
    ws.title = "COMPARISON"
    df = df_orig.copy().reset_index(drop=True)

    orig_cols  = list(df.columns)
    all_cols   = orig_cols + ["NCT PromoSponsor", "Duration +-", "Remarks"]
    total_cols = len(all_cols)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value = "BARC vs NCT Comparison — Original Data with Remarks"
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = CENTER

    # Header row
    for col_i, h in enumerate(all_cols, 1):
        cell = hdr(ws, 2, col_i, h)
        if h == "NCT PromoSponsor":
            cell.fill = PatternFill("solid", fgColor="4A148C")   # dark purple header

    # ── Build a lookup: positional-index-in-NCT-subset → PS label ────────────
    # nct_ps_label keys are positional indices within df_nct (reset_index).
    # We need to map those back to absolute row positions in df_orig.
    nct_positions = df[df["source"].apply(safe_str).str.upper() == "NCT"].index.tolist()
    # nct_positions[i] = absolute iloc in df_orig for the i-th NCT row

    barc_ptr = 0
    nct_ptr  = 0

    for row_i, (_, orig_row) in enumerate(df.iterrows(), 3):
        src = safe_str(orig_row.get("source", "")).upper()

        if src == "BARC XML":
            dur_pm = barc_dur_pm[barc_ptr] if barc_ptr < len(barc_dur_pm) else ""
            remark = barc_remarks[barc_ptr] if barc_ptr < len(barc_remarks) else ""
            ps_col = ""
            barc_ptr += 1
        else:
            dur_pm, remark = "", ""
            # Check if this NCT row was PS-matched
            ps_col = nct_ps_label.get(nct_ptr, "")
            nct_ptr += 1

        # Row fill based on remark
        if src == "BARC XML":
            if remark in ("OK", ""):
                row_fill = OK_FILL
            elif "MISSING" in remark or "WRONG" in remark:
                row_fill = ERR_FILL
            else:
                row_fill = WARN_FILL
        else:
            row_fill = PS_FILL if ps_col else None

        # Write original columns
        for col_i, col_name in enumerate(orig_cols, 1):
            val = safe_str(df.at[orig_row.name, col_name]) if col_name in df.columns else ""
            sc(ws, row_i, col_i, val, fill=row_fill,
               align=LEFT if col_i > 3 else CENTER)

        # NCT PromoSponsor column
        ps_fill = PS_FILL if ps_col else row_fill
        sc(ws, row_i, len(orig_cols)+1, ps_col,
           fill=ps_fill, font=PS_FONT if ps_col else NORMAL_FONT, align=LEFT)
        # Duration +/-
        sc(ws, row_i, len(orig_cols)+2, dur_pm, fill=row_fill, align=CENTER)
        # Remarks
        sc(ws, row_i, len(orig_cols)+3, remark,  fill=row_fill, align=LEFT)

    named_w = {
        "source": 12, "channel name": 20, "TelecastDate": 14,
        "BARC ContentType": 16, "NCT Program Type": 16, "BARC Title": 36,
        "NCT brand": 30, "NCT story": 40, "BARC PromoSponsorName": 36,
        "TelecastStartTime": 14, "TelecastEndTime": 14, "TelecastDuration": 14,
        "NCT PromoSponsor": 30, "Duration +-": 14, "Remarks": 50,
    }
    for col_i, col_name in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = named_w.get(col_name, 14)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

# ── SUMMARY sheet ─────────────────────────────────────────────────────────────

def build_summary_sheet(ws, df_barc, df_nct, brand_matches,
                        barc_brands, ps_brand_nct_map, ps_matched_nct, threshold):
    ws.title = "SUMMARY"

    channel = safe_str(df_barc["channel name"].iloc[0]) if len(df_barc) else ""
    date    = safe_str(df_barc["TelecastDate"].iloc[0])  if len(df_barc) else ""

    def barc_total(ctype):
        return sum(hms_to_secs(v) for v in
                   df_barc[df_barc["BARC ContentType"].apply(safe_str)==ctype]["TelecastDuration"])

    # NCT totals — exclude ps_matched_nct indices from COMMERCIAL counts
    nct_reset = df_nct.copy().reset_index(drop=True)

    def nct_total(ptype):
        mask = (nct_reset["NCT Program Type"].apply(
                    lambda v: normalise_type(safe_str(v))) == ptype) & \
               (~nct_reset.index.isin(ps_matched_nct))
        return sum(hms_to_secs(v) for v in nct_reset[mask]["TelecastDuration"])

    def nct_count(ptype):
        mask = (nct_reset["NCT Program Type"].apply(
                    lambda v: normalise_type(safe_str(v))) == ptype) & \
               (~nct_reset.index.isin(ps_matched_nct))
        return int(mask.sum())

    b_comm  = barc_total("Commercial")
    b_promo = barc_total("Promo")
    b_prog  = barc_total("Program")
    b_ps    = sum(v["secs"] for v in barc_brands.values())

    n_comm  = nct_total("COMMERCIAL")
    n_promo = nct_total("PROMO")
    n_prog  = nct_total("PROGRAM")
    n_ps    = sum(v["secs"]  for v in ps_brand_nct_map.values())
    n_ps_cnt = sum(v["count"] for v in ps_brand_nct_map.values())

    barc_unique = len(barc_brands)
    nct_unique  = len(ps_brand_nct_map)
    nct_matched = sum(1 for m in brand_matches if m["matched"])

    def barc_count(ctype):
        return len(df_barc[df_barc["BARC ContentType"].apply(safe_str) == ctype])

    b_comm_cnt  = barc_count("Commercial")
    b_promo_cnt = barc_count("Promo")
    b_prog_cnt  = barc_count("Program")
    b_ps_cnt    = sum(v["count"] for v in barc_brands.values())

    n_comm_cnt  = nct_count("COMMERCIAL")
    n_promo_cnt = nct_count("PROMO")
    n_prog_cnt  = nct_count("PROGRAM")

    # ══ TABLE 1 ══════════════════════════════════════════════════════════════
    T1R, T1C, T1W = 1, 1, 9
    ws.merge_cells(start_row=T1R, start_column=T1C, end_row=T1R, end_column=T1C+T1W-1)
    ci = ws.cell(row=T1R, column=T1C, value=f"CHANNEL: {channel}    DATE: {date}")
    ci.font = WHITE_FONT; ci.fill = TITLE_FILL; ci.alignment = CENTER

    merge_title(ws, T1R+1, T1C, T1C+T1W-1,
                "TABLE 1 — CONTENT TYPE SUMMARY (BARC vs NCT)", TITLE_FILL)

    for ci, h in enumerate(["SOURCE","CONTENT / PROGRAM TYPE","TOTAL COUNT","TOTAL DURATION",
                             "+- COUNT","+- DURATION",
                             "PROMOSPONSOR UNIQUE COUNT","PROMOSPONSOR TOTAL DURATION",
                             "CONCLUSION / REMARKS"], T1C):
        hdr(ws, T1R+2, ci, h)

    t1_data = [
        ("BARC XML","Commercial",  b_comm_cnt, b_comm, n_comm_cnt, n_comm, "",          ""),
        ("BARC XML","Promo",       b_promo_cnt,b_promo,n_promo_cnt,n_promo,"",          ""),
        ("BARC XML","Program",     b_prog_cnt, b_prog, n_prog_cnt, n_prog, "",          ""),
        ("BARC XML","PromoSponsor",b_ps_cnt,   b_ps,   n_ps_cnt,   n_ps,   barc_unique, b_ps),
        ("NCT",     "COMMERCIAL",  n_comm_cnt, n_comm, b_comm_cnt, b_comm, "",          ""),
        ("NCT",     "PROMO",       n_promo_cnt,n_promo,b_promo_cnt,b_promo,"",          ""),
        ("NCT",     "PROGRAM",     n_prog_cnt, n_prog, b_prog_cnt, b_prog, "",          ""),
        ("NCT",     "PROMOSPONSOR",n_ps_cnt,   n_ps,   b_ps_cnt,   b_ps,   nct_matched, n_ps),
    ]

    row = T1R + 3
    for (src, ctype, own_cnt, own_secs, other_cnt, other_secs, ps_uniq, ps_dur) in t1_data:
        cnt_diff = own_cnt  - other_cnt
        dur_diff = own_secs - other_secs
        label    = "PROMOSPONSOR" if "SPONSOR" in ctype.upper() else ctype.upper()
        cnt_str  = f"{cnt_diff:+d}"      if cnt_diff != 0 else "+0"
        dur_str  = secs_to_hms(dur_diff) if dur_diff != 0 else "+00:00:00"
        other    = "NCT" if src == "BARC XML" else "BARC XML"
        if own_secs == other_secs:
            conc = f"{src} AND {other} {label} DURATION MATCHES PERFECTLY"
        elif dur_diff > 0:
            conc = f"{src} IS TAGGING {fmt_hms(abs(dur_diff))} MORE {label} DURATION THAN {other}"
        else:
            conc = f"{src} IS TAGGING {fmt_hms(abs(dur_diff))} LESS {label} DURATION THAN {other}"

        fill = BARC_FILL if src == "BARC XML" else NCT_FILL
        sc(ws, row, T1C,   src,                                         fill=fill,      align=CENTER)
        sc(ws, row, T1C+1, ctype,                                        fill=fill,      align=CENTER)
        sc(ws, row, T1C+2, own_cnt,                                      fill=fill,      align=CENTER)
        sc(ws, row, T1C+3, fmt_hms(own_secs),                            fill=fill,      align=CENTER)
        sc(ws, row, T1C+4, cnt_str, fill=WARN_FILL if cnt_diff!=0 else OK_FILL,          align=CENTER)
        sc(ws, row, T1C+5, dur_str, fill=WARN_FILL if dur_diff!=0 else OK_FILL,          align=CENTER)
        sc(ws, row, T1C+6, ps_uniq if ps_uniq != "" else "—",            fill=fill,      align=CENTER)
        sc(ws, row, T1C+7, fmt_hms(ps_dur) if ps_dur != "" else "—",     fill=fill,      align=CENTER)
        sc(ws, row, T1C+8, conc,                                          fill=CONC_FILL, align=LEFT,
           font=CONC_FONT)
        row += 1

    # ══ TABLE 2 — PS brand detail ════════════════════════════════════════════
    T2R = row + 2
    merge_title(ws, T2R, 1, 9,
                "TABLE 2 — PROMOSPONSOR BRAND DETAIL (BARC Brand vs NCT Brand)", TITLE_FILL)
    for ci, h in enumerate(["#","BARC BRAND","BARC COUNT","BARC DURATION",
                             "NCT BRAND","NCT COUNT","NCT DURATION",
                             "DURATION +/-","STATUS"], 1):
        hdr(ws, T2R+1, ci, h)

    ba_tot_cnt = ba_tot_secs = an_tot_cnt = an_tot_secs = 0
    for i, m in enumerate(brand_matches, 1):
        fill  = MATCH_FILL if m["matched"] else MISS_FILL
        dur_d = secs_to_hms(m["dur_diff"]) if m["dur_diff"] is not None else "—"
        n_cnt = m["nct_count"] if m["matched"] else "—"
        n_dur = fmt_hms(m["nct_secs"]) if m["matched"] else "—"

        sc(ws, T2R+1+i, 1, i,                       fill=fill,                              align=CENTER)
        sc(ws, T2R+1+i, 2, m["barc_brand"],          fill=fill,                              align=LEFT)
        sc(ws, T2R+1+i, 3, m["barc_count"],          fill=BARC_FILL,                         align=CENTER)
        sc(ws, T2R+1+i, 4, fmt_hms(m["barc_secs"]), fill=BARC_FILL,                         align=CENTER)
        sc(ws, T2R+1+i, 5, m["nct_brand"],           fill=fill,                              align=LEFT)
        sc(ws, T2R+1+i, 6, n_cnt, fill=NCT_FILL if m["matched"] else fill,                  align=CENTER)
        sc(ws, T2R+1+i, 7, n_dur, fill=NCT_FILL if m["matched"] else fill,                  align=CENTER)
        sc(ws, T2R+1+i, 8, dur_d,                    fill=fill,                              align=CENTER)
        sc(ws, T2R+1+i, 9, "✓ MATCHED" if m["matched"] else "✗ NOT FOUND",
           fill=fill, font=BOLD_FONT,                                                        align=CENTER)

        ba_tot_cnt  += m["barc_count"]; ba_tot_secs += m["barc_secs"]
        if m["matched"]:
            an_tot_cnt  += m["nct_count"]; an_tot_secs += m["nct_secs"]

    tot = T2R + 1 + len(brand_matches) + 1
    for ci, v in enumerate([
        "TOTAL", f"{barc_unique} brands", ba_tot_cnt, fmt_hms(ba_tot_secs),
        f"{nct_matched} matched", an_tot_cnt, fmt_hms(an_tot_secs),
        secs_to_hms(an_tot_secs - ba_tot_secs),
        f"{nct_matched}/{barc_unique} BRANDS FOUND IN NCT"
    ], 1):
        sc(ws, tot, ci, v, fill=HDR_FILL, font=WHITE_FONT, align=CENTER)

    # ══ TABLE 3 ══════════════════════════════════════════════════════════════
    T3R = tot + 3
    merge_title(ws, T3R, 1, 9, "TABLE 3 — UNIQUE BRAND COUNT SUMMARY", TITLE_FILL)
    for ci, h in enumerate(["METRIC","BARC XML","NCT","DIFFERENCE","CONCLUSION"], 1):
        hdr(ws, T3R+1, ci, h)

    for i, (metric, bv, nv, diff, conc) in enumerate([
        ("Unique Brands (PromoSponsor)", barc_unique, nct_unique,
         nct_unique - barc_unique,
         f"NCT HAS {abs(nct_unique-barc_unique)} "
         f"{'MORE' if nct_unique>barc_unique else 'FEWER'} UNIQUE BRANDS THAN BARC"),
        ("Brands Matched in NCT", barc_unique, nct_matched, nct_matched - barc_unique,
         f"{nct_matched} OF {barc_unique} BARC BRANDS FOUND IN NCT"
         + (f" | {barc_unique-nct_matched} MISSING" if nct_matched < barc_unique
            else " | ALL MATCHED")),
        ("PromoSponsor Total Duration", b_ps, n_ps, n_ps - b_ps,
         generate_conclusion("PROMOSPONSOR", b_ps, n_ps)),
    ], 1):
        fill = OK_FILL if diff == 0 else (WARN_FILL if diff > 0 else ERR_FILL)
        sc(ws, T3R+1+i, 1, metric, fill=fill, align=LEFT)
        sc(ws, T3R+1+i, 2, fmt_hms(bv) if isinstance(bv, int) and bv > 100 else bv,
           fill=BARC_FILL, align=CENTER)
        sc(ws, T3R+1+i, 3, fmt_hms(nv) if isinstance(nv, int) and nv > 100 else nv,
           fill=NCT_FILL, align=CENTER)
        sc(ws, T3R+1+i, 4,
           secs_to_hms(diff) if isinstance(diff, int) and abs(diff) > 100
           else (f"{diff:+d}" if diff != 0 else "0"),
           fill=fill, align=CENTER)
        sc(ws, T3R+1+i, 5, conc, fill=CONC_FILL, font=CONC_FONT, align=LEFT)

    for col, w in {1:12,2:30,3:12,4:18,5:30,6:12,7:18,8:14,9:50}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20; ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A4"

# ── COMMERCIAL COMPARISON sheet ───────────────────────────────────────────────

def build_commercial_sheet(ws, df_barc, df_nct, threshold, ps_matched_nct,
                           ps_brand_nct_map):
    """
    Top table:
      For each BARC commercial title, show:
        BARC count | BARC duration |
        NCT commercial count (PS rows excluded) | NCT commercial duration (PS rows excluded) |
        NCT PS count | NCT PS duration |
        REMARKS

    A brand like PRAYAG PTMT FAUCETS may have:
      - 26 pure NCT commercial rows  → shown in NCT COMMERCIAL columns
      -  1 NCT PS row               → shown in NCT PS columns

    Bottom table:
      All NCT commercial brands not matched to any BARC title (PS rows never appear here).
    """
    ws.title = "COMMERCIAL COMPARISION"

    channel = safe_str(df_barc["channel name"].iloc[0]) if len(df_barc) else ""
    date    = safe_str(df_barc["TelecastDate"].iloc[0])  if len(df_barc) else ""
    hms_s   = lambda v: hms_to_secs(safe_str(v))

    # ── Aggregate BARC Commercial ─────────────────────────────────────────────
    barc_comm_df = df_barc[df_barc["BARC ContentType"].apply(safe_str) == "Commercial"]
    barc_agg = {}
    for _, row in barc_comm_df.iterrows():
        title = safe_str(row.get("BARC Title", "")) or "—"
        secs  = hms_s(row.get("TelecastDuration", ""))
        e = barc_agg.setdefault(title, {"count": 0, "secs": 0})
        e["count"] += 1; e["secs"] += secs

    # ── Aggregate NCT Commercial — split into PURE commercial vs PS ───────────
    # For each NCT brand, build two buckets:
    #   "comm"  — rows NOT in ps_matched_nct
    #   "ps"    — rows IN ps_matched_nct
    nct_reset = df_nct.reset_index(drop=True)
    nct_comm_pure = {}   # brand -> {count, secs}  — only non-PS commercial rows
    nct_comm_ps   = {}   # brand -> {count, secs}  — only PS-matched commercial rows

    for pos_idx, row in nct_reset.iterrows():
        ptype = normalise_type(safe_str(row.get("NCT Program Type", "")))
        if ptype != "COMMERCIAL":
            continue
        brand = safe_str(row.get("NCT brand", ""))
        if not brand or brand.upper() in INVALID_BRAND_NAMES:
            continue
        secs = hms_s(row.get("TelecastDuration", ""))

        if pos_idx in ps_matched_nct:
            # This is a PS row — but which BARC brand does it belong to?
            # Match against ps_brand_nct_map keys using smart_match
            # We use the brand name from the NCT row itself to find the right bucket
            best_key, best_score = None, 0.0
            for ps_key in ps_brand_nct_map:
                sc_val = smart_match(ps_key, brand)
                if sc_val > best_score:
                    best_score, best_key = sc_val, ps_key
            bucket_key = best_key if best_key and best_score >= threshold else brand
            e = nct_comm_ps.setdefault(bucket_key, {"count": 0, "secs": 0})
            e["count"] += 1; e["secs"] += secs
        else:
            e = nct_comm_pure.setdefault(brand, {"count": 0, "secs": 0})
            e["count"] += 1; e["secs"] += secs

    nct_pure_brands = list(nct_comm_pure.keys())

    # ── Match each BARC commercial → best pure NCT COMMERCIAL brand ───────────
    matched_nct_brands = set()
    section_a = []

    for barc_title in sorted(barc_agg.keys()):
        b_info = barc_agg[barc_title]

        # Find best match in pure NCT commercial brands
        best_nct, _ = best_match_in_list(barc_title, nct_pure_brands, threshold)

        # Check if this BARC title also appears in PS (same brand in PS bucket)
        ps_info = None
        for ps_key in nct_comm_ps:
            if smart_match(barc_title, ps_key) >= threshold:
                ps_info = nct_comm_ps[ps_key]
                break
        # Also check ps_brand_nct_map directly
        if ps_info is None:
            for ps_key, ps_val in ps_brand_nct_map.items():
                if smart_match(barc_title, ps_key) >= threshold:
                    # Look up in nct_comm_ps
                    for ck, cv in nct_comm_ps.items():
                        if smart_match(ps_key, ck) >= threshold:
                            ps_info = cv
                            break
                    if ps_info is None:
                        ps_info = {"count": 0, "secs": 0}
                    break

        if best_nct:
            matched_nct_brands.add(best_nct)
            n_info = nct_comm_pure[best_nct]
            section_a.append({
                "barc_title":  barc_title,
                "b_count":     b_info["count"],  "b_secs":     b_info["secs"],
                "nct_brand":   best_nct,
                "nct_c_count": n_info["count"],  "nct_c_secs": n_info["secs"],
                "nct_ps_count": (ps_info["count"] if ps_info else 0),
                "nct_ps_secs":  (ps_info["secs"]  if ps_info else 0),
                "remark": "MATCHED", "fill_r": MATCH_FILL, "fill_n": NCT_FILL,
            })
        else:
            section_a.append({
                "barc_title":  barc_title,
                "b_count":     b_info["count"],  "b_secs":     b_info["secs"],
                "nct_brand":   "—",
                "nct_c_count": "—",              "nct_c_secs": 0,
                "nct_ps_count": (ps_info["count"] if ps_info else 0),
                "nct_ps_secs":  (ps_info["secs"]  if ps_info else 0),
                "remark": "NOT FOUND IN NCT", "fill_r": MISS_FILL, "fill_n": MISS_FILL,
            })

    # ── Bottom table: unmatched pure NCT commercial brands ────────────────────
    section_b = []
    for nct_brand in sorted(nct_comm_pure.keys()):
        if nct_brand in matched_nct_brands:
            continue
        n_info = nct_comm_pure[nct_brand]
        section_b.append({
            "nct_brand": nct_brand,
            "nct_count": n_info["count"],
            "nct_secs":  n_info["secs"],
        })

    # ── Write sheet ───────────────────────────────────────────────────────────
    COLS = 12
    headers = [
        "SOURCE", "CHANNEL NAME", "DATE",
        "BARC COMMERCIAL CONTENT NAME", "NCT COMMERCIAL CONTENT NAME",
        "BARC COUNT", "NCT COMMERCIAL COUNT", "BARC DURATION", "NCT COMMERCIAL DURATION",
        "NCT PS COUNT", "NCT PS DURATION",
        "REMARKS",
    ]

    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    c = ws["A1"]
    c.value = f"COMMERCIAL COMPARISION — {channel}  |  {date}"
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = CENTER

    for ci, h in enumerate(headers, 1):
        cell = hdr(ws, 2, ci, h)
        if "PS" in h:
            cell.fill = PatternFill("solid", fgColor="4A148C")

    cur = 3
    merge_title(ws, cur, 1, COLS, "BARC COMMERCIAL vs NCT COMMERCIAL — MATCHED", TITLE_FILL)
    cur += 1

    ba_tot_cnt = ba_tot_secs = an_tot_cnt = an_tot_secs = 0
    an_ps_cnt  = an_ps_secs  = 0

    for r in section_a:
        nc = r["nct_c_count"]; nd = fmt_hms(r["nct_c_secs"]) if isinstance(r["nct_c_secs"], int) and r["nct_c_secs"] > 0 else "—"
        ps_c = r["nct_ps_count"]; ps_d = fmt_hms(r["nct_ps_secs"]) if r["nct_ps_secs"] > 0 else "—"
        sc(ws, cur,  1, "BARC XML",           fill=r["fill_r"], align=CENTER)
        sc(ws, cur,  2, channel,               fill=r["fill_r"], align=CENTER)
        sc(ws, cur,  3, date,                  fill=r["fill_r"], align=CENTER)
        sc(ws, cur,  4, r["barc_title"],       fill=BARC_FILL,   align=LEFT)
        sc(ws, cur,  5, r["nct_brand"],        fill=r["fill_n"], align=LEFT)
        sc(ws, cur,  6, r["b_count"],          fill=BARC_FILL,   align=CENTER)
        sc(ws, cur,  7, nc,                    fill=r["fill_n"], align=CENTER)
        sc(ws, cur,  8, fmt_hms(r["b_secs"]), fill=BARC_FILL,   align=CENTER)
        sc(ws, cur,  9, nd,                    fill=r["fill_n"], align=CENTER)
        sc(ws, cur, 10, ps_c if ps_c else "—", fill=PS_FILL,    align=CENTER,
           font=PS_FONT if ps_c else NORMAL_FONT)
        sc(ws, cur, 11, ps_d,                  fill=PS_FILL,    align=CENTER,
           font=PS_FONT if r["nct_ps_secs"] > 0 else NORMAL_FONT)
        sc(ws, cur, 12, r["remark"],           fill=r["fill_r"], align=LEFT, font=BOLD_FONT)

        ba_tot_cnt  += r["b_count"];  ba_tot_secs += r["b_secs"]
        if isinstance(nc, int):
            an_tot_cnt += nc; an_tot_secs += r["nct_c_secs"]
        if isinstance(ps_c, int):
            an_ps_cnt += ps_c; an_ps_secs += r["nct_ps_secs"]
        cur += 1

    # Matching total row
    matched_count = sum(1 for r in section_a if r["remark"] == "MATCHED")
    for ci in range(1, COLS+1): sc(ws, cur, ci, "", fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  1, "MATCHING COMMERCIAL TOTAL",    fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  4, f"{len(barc_agg)} BARC unique", fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  5, f"{matched_count} NCT matched", fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  6, ba_tot_cnt,                     fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  7, an_tot_cnt,                     fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  8, fmt_hms(ba_tot_secs),           fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  9, fmt_hms(an_tot_secs),           fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur, 10, an_ps_cnt,                      fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur, 11, fmt_hms(an_ps_secs),            fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    cur += 2

    # Bottom table — unmatched NCT commercials
    merge_title(ws, cur, 1, COLS,
                "NCT COMMERCIAL BRANDS — NOT MATCHED TO ANY BARC COMMERCIAL", SEC_FILL)
    cur += 1

    nb_tot_cnt = nb_tot_secs = 0
    if section_b:
        for r in section_b:
            sc(ws, cur,  1, "NCT",                 fill=ERR_FILL, align=CENTER)
            sc(ws, cur,  2, channel,                fill=ERR_FILL, align=CENTER)
            sc(ws, cur,  3, date,                   fill=ERR_FILL, align=CENTER)
            sc(ws, cur,  4, "—",                    fill=ERR_FILL, align=CENTER)
            sc(ws, cur,  5, r["nct_brand"],         fill=NCT_FILL, align=LEFT)
            sc(ws, cur,  6, "—",                    fill=ERR_FILL, align=CENTER)
            sc(ws, cur,  7, r["nct_count"],         fill=NCT_FILL, align=CENTER)
            sc(ws, cur,  8, "—",                    fill=ERR_FILL, align=CENTER)
            sc(ws, cur,  9, fmt_hms(r["nct_secs"]),fill=NCT_FILL, align=CENTER)
            sc(ws, cur, 10, "—",                    fill=ERR_FILL, align=CENTER)
            sc(ws, cur, 11, "—",                    fill=ERR_FILL, align=CENTER)
            sc(ws, cur, 12, "NOT MATCHED TO ANY BARC COMMERCIAL",
               fill=ERR_FILL, align=LEFT, font=BOLD_FONT)
            nb_tot_cnt += r["nct_count"]; nb_tot_secs += r["nct_secs"]
            cur += 1
    else:
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=COLS)
        c = ws.cell(row=cur, column=1, value="ALL NCT COMMERCIAL BRANDS MATCHED TO BARC")
        c.font = BOLD_FONT; c.fill = OK_FILL; c.alignment = CENTER; c.border = THIN_BORDER
        cur += 1

    # Unmatched total row
    for ci in range(1, COLS+1): sc(ws, cur, ci, "", fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  1, "NCT UNMATCHED TOTAL",          fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  5, f"{len(section_b)} unmatched",  fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  7, nb_tot_cnt,                     fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  9, fmt_hms(nb_tot_secs),           fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    cur += 2

    # Grand Total row
    nct_pure_total_cnt  = sum(v["count"] for v in nct_comm_pure.values())
    nct_pure_total_secs = sum(v["secs"]  for v in nct_comm_pure.values())
    nct_ps_total_cnt    = sum(v["count"] for v in nct_comm_ps.values())
    nct_ps_total_secs   = sum(v["secs"]  for v in nct_comm_ps.values())
    for ci in range(1, COLS+1): sc(ws, cur, ci, "", fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  1, "GRAND TOTAL",                    fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  4, f"{len(barc_agg)} BARC unique",   fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  5, f"{len(nct_comm_pure)} NCT unique",fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  6, ba_tot_cnt,                        fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  7, nct_pure_total_cnt,                fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  8, fmt_hms(ba_tot_secs),              fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur,  9, fmt_hms(nct_pure_total_secs),      fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur, 10, nct_ps_total_cnt,                  fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    sc(ws, cur, 11, fmt_hms(nct_ps_total_secs),        fill=HDR_FILL, font=WHITE_FONT, align=CENTER)
    conc = generate_conclusion("COMMERCIAL", ba_tot_secs, nct_pure_total_secs)
    sc(ws, cur, 12, conc, fill=CONC_FILL, font=CONC_FONT, align=LEFT)

    col_w = {1:12,2:16,3:12,4:36,5:36,6:10,7:14,8:14,9:16,10:12,11:14,12:50}
    for col, w in col_w.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

# ── Summary — add commercial summary ─────────────────────────────────────────

def add_commercial_summary_to_summary(ws, df_barc, df_nct, ps_matched_nct):
    hms_s = lambda v: hms_to_secs(safe_str(v))

    barc_comm       = df_barc[df_barc["BARC ContentType"].apply(safe_str) == "Commercial"]
    barc_unique     = barc_comm["BARC Title"].apply(safe_str).replace("", "—").nunique()
    barc_total_cnt  = len(barc_comm)
    barc_total_secs = sum(hms_s(v) for v in barc_comm["TelecastDuration"])

    nct_reset = df_nct.reset_index(drop=True)
    # Pure NCT commercial = COMMERCIAL type AND NOT in ps_matched_nct
    nct_comm = nct_reset[
        (nct_reset["NCT Program Type"].apply(lambda v: normalise_type(safe_str(v))) == "COMMERCIAL") &
        (~nct_reset.index.isin(ps_matched_nct))
    ]
    valid_brands    = nct_comm["NCT brand"].apply(safe_str)
    valid_brands    = valid_brands[(valid_brands != "") &
                                   (~valid_brands.str.upper().isin(INVALID_BRAND_NAMES))]
    nct_unique      = valid_brands.nunique()
    nct_total_cnt   = len(nct_comm)
    nct_total_secs  = sum(hms_s(v) for v in nct_comm["TelecastDuration"])

    last = ws.max_row + 3
    merge_title(ws, last, 1, 5, "COMMERCIAL MATCH SUMMARY", TITLE_FILL)
    last += 1

    for ci, h in enumerate(["SOURCE","UNIQUE COUNT","TOTAL COUNT","TOTAL DURATION",""], 1):
        hdr(ws, last, ci, h)
    last += 1

    for ci, v in enumerate(["BARC XML", barc_unique, barc_total_cnt,
                             fmt_hms(barc_total_secs), ""], 1):
        sc(ws, last, ci, v, fill=BARC_FILL, align=CENTER)
    last += 1
    for ci, v in enumerate(["NCT", nct_unique, nct_total_cnt,
                             fmt_hms(nct_total_secs), ""], 1):
        sc(ws, last, ci, v, fill=NCT_FILL, align=CENTER)
    last += 2

    cnt_diff = nct_total_cnt  - barc_total_cnt
    dur_diff = nct_total_secs - barc_total_secs
    conc = (f"NCT IS {abs(cnt_diff)} COUNT {'MORE' if cnt_diff >= 0 else 'LESS'} AND "
            f"{fmt_hms(abs(dur_diff))} DURATION {'MORE' if dur_diff >= 0 else 'LESS'} "
            f"THAN BARC IN COMMERCIAL")
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=5)
    c = ws.cell(row=last, column=1, value=conc)
    c.font = CONC_FONT; c.fill = CONC_FILL; c.alignment = LEFT; c.border = THIN_BORDER

# ── TABSONS SUMMARY sheet ────────────────────────────────────────────────────

def build_tabsons_summary_sheet(ws, df_barc, df_nct,
                                barc_brands, ps_brand_nct_map, ps_matched_nct):
    """
    Single-row summary sheet with 37 columns matching the Tabsons reporting format.
    Tabsons = NCT (same data, different name).
    ICA columns are always 0 as per business requirement.
    """
    ws.title = "TABSONS SUMMARY"

    # ── Pull channel / date ───────────────────────────────────────────────────
    channel = safe_str(df_barc["channel name"].iloc[0]) if len(df_barc) else ""
    date    = safe_str(df_barc["TelecastDate"].iloc[0])  if len(df_barc) else ""

    hms_s = lambda v: hms_to_secs(safe_str(v))

    # ── BARC totals ───────────────────────────────────────────────────────────
    barc_total_cnt  = len(df_barc)
    barc_total_secs = sum(hms_s(v) for v in df_barc["TelecastDuration"])

    def barc_cnt(ctype):
        return len(df_barc[df_barc["BARC ContentType"].apply(safe_str) == ctype])
    def barc_dur(ctype):
        return sum(hms_s(v) for v in
                   df_barc[df_barc["BARC ContentType"].apply(safe_str) == ctype]["TelecastDuration"])

    barc_comm_cnt  = barc_cnt("Commercial")
    barc_comm_dur  = barc_dur("Commercial")
    barc_promo_cnt = barc_cnt("Promo")
    barc_promo_dur = barc_dur("Promo")
    barc_prog_cnt  = barc_cnt("Program")
    barc_prog_dur  = barc_dur("Program")

    # BARC PromoSponsor totals (from barc_brands built in analyse_brands)
    barc_ps_cnt  = sum(v["count"] for v in barc_brands.values())
    barc_ps_dur  = sum(v["secs"]  for v in barc_brands.values())

    # BARC unique commercial titles
    barc_unique_comm = df_barc[df_barc["BARC ContentType"].apply(safe_str) == "Commercial"][
        "BARC Title"].apply(safe_str).replace("", "—").nunique()

    # ── NCT (Tabsons) totals ──────────────────────────────────────────────────
    nct_reset = df_nct.copy().reset_index(drop=True)

    nct_total_cnt  = len(nct_reset)
    nct_total_secs = sum(hms_s(v) for v in nct_reset["TelecastDuration"])

    def nct_cnt(ptype, exclude_ps=False):
        mask = nct_reset["NCT Program Type"].apply(
                   lambda v: normalise_type(safe_str(v))) == ptype
        if exclude_ps:
            mask = mask & (~nct_reset.index.isin(ps_matched_nct))
        return int(mask.sum())

    def nct_dur(ptype, exclude_ps=False):
        mask = nct_reset["NCT Program Type"].apply(
                   lambda v: normalise_type(safe_str(v))) == ptype
        if exclude_ps:
            mask = mask & (~nct_reset.index.isin(ps_matched_nct))
        return sum(hms_s(v) for v in nct_reset[mask]["TelecastDuration"])

    # Commercial: exclude PS-matched rows (same logic as existing SUMMARY sheet)
    nct_comm_cnt  = nct_cnt("COMMERCIAL", exclude_ps=True)
    nct_comm_dur  = nct_dur("COMMERCIAL", exclude_ps=True)
    nct_promo_cnt = nct_cnt("PROMO")
    nct_promo_dur = nct_dur("PROMO")
    nct_prog_cnt  = nct_cnt("PROGRAM")
    nct_prog_dur  = nct_dur("PROGRAM")

    # NCT PromoSponsor (from ps_brand_nct_map built in compare_rows)
    nct_ps_cnt  = sum(v["count"] for v in ps_brand_nct_map.values())
    nct_ps_dur  = sum(v["secs"]  for v in ps_brand_nct_map.values())

    # NCT unique commercial brands (excluding PS rows and invalid names)
    nct_comm_pure = nct_reset[
        (nct_reset["NCT Program Type"].apply(
            lambda v: normalise_type(safe_str(v))) == "COMMERCIAL") &
        (~nct_reset.index.isin(ps_matched_nct))
    ]
    valid_nct_brands = nct_comm_pure["NCT brand"].apply(safe_str)
    valid_nct_brands = valid_nct_brands[
        (valid_nct_brands != "") &
        (~valid_nct_brands.str.upper().isin(INVALID_BRAND_NAMES))
    ]
    nct_unique_comm = valid_nct_brands.nunique()

    # ── Accuracy % helpers ────────────────────────────────────────────────────
    def pct(num, den):
        """Return rounded % string, or '0.00%' if denominator is 0."""
        if den == 0:
            return "0.00%"
        return f"{round((num / den) * 100, 2):.2f}%"

    # ── Build the 37 columns ──────────────────────────────────────────────────
    HEADERS = [
        "CHANNEL NAME",
        "CHANNEL DATE",
        "TABSONS LINE ITEM",
        "TABSONS DURATION",
        "BARC LINE ITEM",
        "BARC DURATION",
        "BARC ICA COUNT",
        "BARC ICA DURATION",
        "TABSONS ICA (COUNT)",
        "TABSONS ICA (DURATION)",
        "TABSONS COMMERCIAL COUNT",
        "TABSONS COMMERCIAL DURATION",
        "BARC COMMERCIAL COUNT",
        "BARC COMMERCIAL DURATION",
        "ACCURACY-COMMERCIAL COUNT % (TAB)",
        "ACCURACY DURATION COMMERCIAL % (TAB)",
        "TABSONS PROMO COUNT",
        "TABSONS PROMO DURATION",
        "BARC PROMO COUNT",
        "BARC PROMO DURATION",
        "TABSONS PROMO % (COUNT) ACCURACY",
        "TABSONS PROMO DURATION% (DUR) ACCURACY",
        "BARC PROMO %",
        "BARC PROMO DURATION%",
        "TABSONS PROMO SPONSOR COUNT",
        "TABSONS PROMO SPONSOR COUNT DURATION",
        "BARC PROMO SPONSOR COUNT",
        "BARC PROMO SPONSOR COUNT DURATION",
        "TABSONS PROMO SPONSOR %",
        "TABSONS PROMO SPONSOR DURATION%",
        "TABSONS PROGRAM COUNT",
        "TABSONS PROGRAM DURATION",
        "BARC PROGRAM COUNT",
        "BARC PROGRAM DURATION",
        "TABSONS PROGRAM DURATION ACCURACY%",
        "TABSONS UNIQUE COMMERCIAL COUNT",
        "BARC UNIQUE COMMERCIAL COUNT",
    ]

    VALUES = [
        channel,                                                   # CHANNEL NAME
        date,                                                      # CHANNEL DATE
        nct_total_cnt,                                             # TABSONS LINE ITEM
        fmt_hms(nct_total_secs),                                   # TABSONS DURATION
        barc_total_cnt,                                            # BARC LINE ITEM
        fmt_hms(barc_total_secs),                                  # BARC DURATION
        0,                                                         # BARC ICA COUNT
        fmt_hms(0),                                                # BARC ICA DURATION
        0,                                                         # TABSONS ICA (COUNT)
        fmt_hms(0),                                                # TABSONS ICA (DURATION)
        nct_comm_cnt,                                              # TABSONS COMMERCIAL COUNT
        fmt_hms(nct_comm_dur),                                     # TABSONS COMMERCIAL DURATION
        barc_comm_cnt,                                             # BARC COMMERCIAL COUNT
        fmt_hms(barc_comm_dur),                                    # BARC COMMERCIAL DURATION
        pct(nct_comm_cnt,  barc_comm_cnt),                         # ACCURACY-COMMERCIAL COUNT %
        pct(nct_comm_dur,  barc_comm_dur),                         # ACCURACY DURATION COMMERCIAL %
        nct_promo_cnt,                                             # TABSONS PROMO COUNT
        fmt_hms(nct_promo_dur),                                    # TABSONS PROMO DURATION
        barc_promo_cnt,                                            # BARC PROMO COUNT
        fmt_hms(barc_promo_dur),                                   # BARC PROMO DURATION
        pct(nct_promo_cnt, barc_promo_cnt),                        # TABSONS PROMO % COUNT ACCURACY
        pct(nct_promo_dur, barc_promo_dur),                        # TABSONS PROMO DURATION% ACCURACY
        pct(barc_promo_cnt, barc_total_cnt),                       # BARC PROMO %
        pct(barc_promo_dur, barc_total_secs),                      # BARC PROMO DURATION%
        nct_ps_cnt,                                                # TABSONS PROMO SPONSOR COUNT
        fmt_hms(nct_ps_dur),                                       # TABSONS PROMO SPONSOR DURATION
        barc_ps_cnt,                                               # BARC PROMO SPONSOR COUNT
        fmt_hms(barc_ps_dur),                                      # BARC PROMO SPONSOR DURATION
        pct(nct_ps_cnt,  barc_ps_cnt),                             # TABSONS PROMO SPONSOR %
        pct(nct_ps_dur,  barc_ps_dur),                             # TABSONS PROMO SPONSOR DURATION%
        nct_prog_cnt,                                              # TABSONS PROGRAM COUNT
        fmt_hms(nct_prog_dur),                                     # TABSONS PROGRAM DURATION
        barc_prog_cnt,                                             # BARC PROGRAM COUNT
        fmt_hms(barc_prog_dur),                                    # BARC PROGRAM DURATION
        pct(nct_prog_dur, barc_prog_dur),                          # TABSONS PROGRAM DURATION ACCURACY%
        nct_unique_comm,                                           # TABSONS UNIQUE COMMERCIAL COUNT
        barc_unique_comm,                                          # BARC UNIQUE COMMERCIAL COUNT
    ]

    TOTAL_COLS = len(HEADERS)

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(TOTAL_COLS)}1")
    c = ws["A1"]
    c.value = f"TABSONS SUMMARY — {channel}  |  {date}"
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = CENTER
    ws.row_dimensions[1].height = 22

    # ── Header row (row 2) ────────────────────────────────────────────────────
    # Colour groups for readability
    GROUP_FILLS = {
        # col index (1-based) -> fill
        1:  BARC_FILL,   2:  BARC_FILL,                          # channel info
        3:  NCT_FILL,    4:  NCT_FILL,                           # tabsons totals
        5:  BARC_FILL,   6:  BARC_FILL,                          # barc totals
        7:  WARN_FILL,   8:  WARN_FILL,                          # barc ICA
        9:  WARN_FILL,   10: WARN_FILL,                          # tabsons ICA
        11: NCT_FILL,    12: NCT_FILL,                           # tabsons commercial
        13: BARC_FILL,   14: BARC_FILL,                          # barc commercial
        15: OK_FILL,     16: OK_FILL,                            # commercial accuracy
        17: NCT_FILL,    18: NCT_FILL,                           # tabsons promo
        19: BARC_FILL,   20: BARC_FILL,                          # barc promo
        21: OK_FILL,     22: OK_FILL,                            # promo count/dur accuracy
        23: BARC_FILL,   24: BARC_FILL,                          # barc promo %
        25: PS_FILL,     26: PS_FILL,                            # tabsons PS
        27: BARC_FILL,   28: BARC_FILL,                          # barc PS
        29: PS_FILL,     30: PS_FILL,                            # tabsons PS %
        31: NCT_FILL,    32: NCT_FILL,                           # tabsons program
        33: BARC_FILL,   34: BARC_FILL,                          # barc program
        35: OK_FILL,                                             # program accuracy
        36: NCT_FILL,    37: BARC_FILL,                          # unique commercial
    }

    for col_i, h in enumerate(HEADERS, 1):
        fill = GROUP_FILLS.get(col_i, HDR_FILL)
        # ICA columns: always use WARN fill with darker text to signal "always 0"
        if "ICA" in h:
            fill = WARN_FILL
        cell = ws.cell(row=2, column=col_i, value=h)
        cell.font      = WHITE_FONT if fill == HDR_FILL else Font(name="Arial", bold=True, size=9, color="1F2937")
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = 18
    ws.row_dimensions[2].height = 50

    # ── Data row (row 3) ──────────────────────────────────────────────────────
    for col_i, val in enumerate(VALUES, 1):
        fill = GROUP_FILLS.get(col_i, PatternFill())
        # Accuracy % cells: green if >= 95%, amber if >= 80%, red otherwise
        is_pct = isinstance(val, str) and val.endswith("%")
        if is_pct:
            try:
                pct_val = float(val.replace("%", ""))
                if pct_val >= 95:
                    fill = OK_FILL
                elif pct_val >= 80:
                    fill = WARN_FILL
                else:
                    fill = ERR_FILL
            except ValueError:
                pass
        # ICA cols: always light amber, value 0
        if "ICA" in HEADERS[col_i - 1]:
            fill = PatternFill("solid", fgColor="FFF3CD")
        cell = ws.cell(row=3, column=col_i, value=val)
        cell.font      = NORMAL_FONT
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER

    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A3"


# ── DETAILED ANALYSIS sheet ───────────────────────────────────────────────────

def build_detailed_analysis_sheet(ws, df_barc, df_nct, threshold, tol):
    ws.title = "DETAILED ANALYSIS"

    channel = safe_str(df_barc["channel name"].iloc[0]) if len(df_barc) else ""
    date    = safe_str(df_barc["TelecastDate"].iloc[0])  if len(df_barc) else ""

    nct = df_nct.copy().reset_index(drop=True)
    nct["_ss"] = nct["TelecastStartTime"].apply(hms_to_secs)
    nct["_se"] = nct["TelecastEndTime"].apply(hms_to_secs)

    COLS = 9
    headers = ["BARC CONTENT NAME","BARC CONTENT TYPE","BARC START TIME","BARC END TIME",
               "NCT CONTENT NAME","NCT CONTENT TYPE","NCT START TIME","NCT END TIME","REMARKS"]

    ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
    c = ws["A1"]
    c.value = f"DETAILED ANALYSIS — {channel}  |  {date}"
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = CENTER
    for ci, h in enumerate(headers, 1): hdr(ws, 2, ci, h)

    cur = 3; rows_written = 0

    for _, brow in df_barc.iterrows():
        b_start     = hms_to_secs(brow["TelecastStartTime"])
        b_end       = hms_to_secs(brow["TelecastEndTime"])
        b_ctype     = normalise_type(safe_str(brow["BARC ContentType"]))
        b_title     = safe_str(brow["BARC Title"])
        b_spname    = safe_str(brow["BARC PromoSponsorName"])
        b_start_str = safe_str(brow["TelecastStartTime"])
        b_end_str   = safe_str(brow["TelecastEndTime"])

        if b_ctype not in ("COMMERCIAL", "PROMO"): continue

        is_sp   = (b_ctype == "PROMO" and b_title.lower().startswith("sponsorship promo"))
        mask    = (nct["_ss"] >= (b_start - tol)) & (nct["_ss"] < (b_end + tol))
        nct_win = nct[mask].copy()
        if len(nct_win) == 0: continue

        if is_sp and b_spname:
            for brand, _ in parse_sponsor_brands(b_spname):
                found_idx, _, _ = find_brand_in_nct(brand, nct_win, threshold)
                if found_idx is None:
                    for _, nrow in nct_win.iterrows():
                        nct_name  = (safe_str(nrow.get("NCT brand",""))
                                     or safe_str(nrow.get("NCT story","")))
                        nct_ptype = normalise_type(safe_str(nrow.get("NCT Program Type","")))
                        sc(ws, cur, 1, brand,           fill=BARC_FILL, align=LEFT)
                        sc(ws, cur, 2, "PROMO SPONSOR", fill=BARC_FILL, align=CENTER)
                        sc(ws, cur, 3, b_start_str,     fill=BARC_FILL, align=CENTER)
                        sc(ws, cur, 4, b_end_str,       fill=BARC_FILL, align=CENTER)
                        sc(ws, cur, 5, nct_name,        fill=NCT_FILL,  align=LEFT)
                        sc(ws, cur, 6, nct_ptype,       fill=NCT_FILL,  align=CENTER)
                        sc(ws, cur, 7, safe_str(nrow.get("TelecastStartTime","")),
                           fill=NCT_FILL, align=CENTER)
                        sc(ws, cur, 8, safe_str(nrow.get("TelecastEndTime","")),
                           fill=NCT_FILL, align=CENTER)
                        sc(ws, cur, 9,
                           f"NCT is tagging ({nct_name}) instead of ({brand})",
                           fill=ERR_FILL, align=LEFT, font=BOLD_FONT)
                        cur += 1; rows_written += 1
        else:
            for _, nrow in nct_win.iterrows():
                nct_ptype  = normalise_type(safe_str(nrow.get("NCT Program Type","")))
                nct_brand  = safe_str(nrow.get("NCT brand",""))
                nct_story  = safe_str(nrow.get("NCT story",""))
                nct_name   = nct_brand or nct_story
                type_match = (nct_ptype == b_ctype)
                score_b    = smart_match(b_title, nct_brand) if nct_brand else 0.0
                score_s    = smart_match(b_title, nct_story) if nct_story else 0.0
                if max(score_b, score_s) >= threshold: continue
                fill = WARN_FILL if type_match else ERR_FILL
                sc(ws, cur, 1, b_title,     fill=BARC_FILL, align=LEFT)
                sc(ws, cur, 2, b_ctype,     fill=BARC_FILL, align=CENTER)
                sc(ws, cur, 3, b_start_str, fill=BARC_FILL, align=CENTER)
                sc(ws, cur, 4, b_end_str,   fill=BARC_FILL, align=CENTER)
                sc(ws, cur, 5, nct_name,    fill=NCT_FILL,  align=LEFT)
                sc(ws, cur, 6, nct_ptype,   fill=NCT_FILL,  align=CENTER)
                sc(ws, cur, 7, safe_str(nrow.get("TelecastStartTime","")),
                   fill=NCT_FILL, align=CENTER)
                sc(ws, cur, 8, safe_str(nrow.get("TelecastEndTime","")),
                   fill=NCT_FILL, align=CENTER)
                sc(ws, cur, 9,
                   f"NCT is tagging ({nct_name}) instead of ({b_title})",
                   fill=fill, align=LEFT, font=BOLD_FONT)
                cur += 1; rows_written += 1

    if rows_written == 0:
        ws.merge_cells(f"A3:{get_column_letter(COLS)}3")
        c = ws.cell(row=3, column=1,
                    value="NO MISMATCHES FOUND — ALL CONTENT MATCHES CORRECTLY")
        c.font = BOLD_FONT; c.fill = OK_FILL; c.alignment = CENTER; c.border = THIN_BORDER

    for col, w in {1:40,2:18,3:14,4:14,5:40,6:18,7:14,8:14,9:60}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"
    return rows_written

# ── Library entry point (replaces subprocess execution) ──────────────────────

def run_comparison_from_bytes(file_bytes: bytes) -> bytes:
    """
    Run the full BARC/NCT comparison from raw Excel bytes.
    Returns the output workbook as bytes.
    No file I/O — everything in memory.
    """
    import io as _io
    from openpyxl import Workbook as _Workbook

    df = pd.read_excel(_io.BytesIO(file_bytes), dtype=str)
    df.columns = [c.strip() for c in df.columns]

    for col in ["source", "channel name", "TelecastDate", "TelecastStartTime",
                "TelecastEndTime", "TelecastDuration", "BARC ContentType",
                "BARC Title", "BARC PromoSponsorName", "NCT Program Type",
                "NCT brand", "NCT story"]:
        if col not in df.columns:
            df[col] = ""

    df["source"] = df["source"].apply(safe_str).str.upper().str.strip()
    df_barc = df[df["source"] == "BARC XML"].copy().reset_index(drop=True)
    df_nct  = df[df["source"] == "NCT"].copy().reset_index(drop=True)

    barc_dur_pm, barc_remarks, ps_matched_nct, nct_ps_label, ps_brand_nct_map = compare_rows(
        df_barc, df_nct, SIMILARITY_THRESHOLD, TIME_TOLERANCE_SECS)

    barc_brands, brand_matches = analyse_brands(
        df_barc, ps_brand_nct_map, SIMILARITY_THRESHOLD)

    wb = _Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("COMPARISON")
    build_comparison_sheet(ws1, df, barc_dur_pm, barc_remarks, nct_ps_label)

    ws2 = wb.create_sheet("SUMMARY")
    build_summary_sheet(ws2, df_barc, df_nct, brand_matches,
                        barc_brands, ps_brand_nct_map, ps_matched_nct, SIMILARITY_THRESHOLD)

    ws3 = wb.create_sheet("COMMERCIAL COMPARISION")
    build_commercial_sheet(ws3, df_barc, df_nct, SIMILARITY_THRESHOLD,
                           ps_matched_nct, ps_brand_nct_map)

    ws5 = wb.create_sheet("TABSONS SUMMARY")
    build_tabsons_summary_sheet(ws5, df_barc, df_nct,
                                barc_brands, ps_brand_nct_map, ps_matched_nct)

    ws4 = wb.create_sheet("DETAILED ANALYSIS")
    build_detailed_analysis_sheet(ws4, df_barc, df_nct,
                                  SIMILARITY_THRESHOLD, TIME_TOLERANCE_SECS)

    add_commercial_summary_to_summary(ws2, df_barc, df_nct, ps_matched_nct)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── CLI entry point (backward-compatible file-based usage) ───────────────────
if __name__ == "__main__":
    import sys as _sys
    input_file  = "brand_comparison_template.xlsx"
    output_file = "barc_nct_comparison.xlsx"
    print(f"Reading: {input_file}")
    with open(input_file, "rb") as fh:
        result_bytes = run_comparison_from_bytes(fh.read())
    with open(output_file, "wb") as fh:
        fh.write(result_bytes)
    print(f"Output saved: {output_file}")
